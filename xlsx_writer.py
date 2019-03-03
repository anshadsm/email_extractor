import time

from openpyxl import Workbook


class XlsxWrite:
    def __init__(self):
        timestr = time.strftime("%Y%m%d-%H%M%S")
        self.file_name = "Emails-" + timestr + '.xlsx'
        self.file_path = "output/"
        self.file = self.file_path + self.file_name
        self.last_handled_row = 2
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Emails"
        self.create_header()

    def create_header(self):
        self.ws['A1'] = 'No'
        self.ws['B1'] = 'Emails'

    def write_to_sheet(self, items=None):
        for item in items:
            self.ws.cell(self.last_handled_row, 1).value = self.last_handled_row
            self.ws.cell(self.last_handled_row, 2).value = item
            self.last_handled_row += 1

    def save_file(self):
        self.wb.save(self.file)
